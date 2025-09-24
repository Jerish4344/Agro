"""
Views for reporting and analytics.
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib import messages
from django.db.models import Count, Sum, Avg, Min, Max
import csv
import json
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..models import PriceSubmission, Firm, ProductCategory, Product, Location
from ..services.reporting_service import ReportingService
from ..forms import PriceSubmissionFilterForm


@login_required
def reports_index(request):
    """Main reports landing page"""
    user_profile = request.user.profile
    
    # Get quick stats for the dashboard
    today = timezone.now().date()
    
    # Base queryset with user permissions
    submissions_qs = PriceSubmission.objects.filter(status='APPROVED')
    
    if not user_profile.is_admin():
        if user_profile.firm:
            submissions_qs = submissions_qs.filter(firm=user_profile.firm)
        
        if user_profile.is_category_head() and hasattr(user_profile, 'category_head_profile'):
            scoped_categories = user_profile.category_head_profile.scoped_categories.all()
            if scoped_categories:
                submissions_qs = submissions_qs.filter(category__in=scoped_categories)
    
    # Calculate quick stats
    today_stats = submissions_qs.filter(date=today).aggregate(
        total_submissions=Count('id'),
        total_value=Sum('price_per_unit'),
        avg_price=Avg('price_per_unit'),
        min_price=Min('price_per_unit'),
        max_price=Max('price_per_unit')
    )
    
    # Weekly stats
    week_ago = today - timedelta(days=7)
    weekly_stats = submissions_qs.filter(date__gte=week_ago).aggregate(
        total_submissions=Count('id'),
        unique_products=Count('product', distinct=True),
        unique_locations=Count('location', distinct=True)
    )
    
    # Top products by submission count this week
    top_products = submissions_qs.filter(
        date__gte=week_ago
    ).values(
        'product__name', 'product__category__name'
    ).annotate(
        submission_count=Count('id'),
        avg_price=Avg('price_per_unit')
    ).order_by('-submission_count')[:5]
    
    context = {
        'user_profile': user_profile,
        'today_stats': today_stats,
        'weekly_stats': weekly_stats,
        'top_products': top_products,
        'report_date': today
    }
    
    return render(request, 'core/reports/index.html', context)


@login_required
def firm_daily_report(request):
    """Firm-wise daily report"""
    user_profile = request.user.profile
    
    # Get parameters
    date_str = request.GET.get('date', timezone.now().date().isoformat())
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        report_date = timezone.now().date()
    
    firm_id = request.GET.get('firm')
    category_id = request.GET.get('category')
    product_id = request.GET.get('product')
    location_id = request.GET.get('location')
    export_format = request.GET.get('format')
    
    # Get firm filter
    firm = None
    if firm_id:
        try:
            firm = Firm.objects.get(id=firm_id)
        except Firm.DoesNotExist:
            firm = None
    
    # For non-admin users, default to their firm
    if not user_profile.is_admin() and user_profile.firm:
        firm = user_profile.firm
    
    # Get the report data
    if user_profile.is_admin():
        firm_reports = ReportingService.get_firm_wise_daily_report(date=report_date)
        
        # Filter to specific firm if requested
        if firm:
            firm_reports = [report for report in firm_reports if report['firm'] == firm]
    else:
        # Non-admin users get report for their firm only
        firm_reports = ReportingService.get_firm_wise_daily_report(date=report_date)
        if user_profile.firm:
            firm_reports = [report for report in firm_reports if report['firm'] == user_profile.firm]
    
    # Get detailed submissions for the selected criteria
    submissions_qs = PriceSubmission.objects.filter(
        date=report_date,
        status='APPROVED'
    ).select_related(
        'firm', 'product', 'category', 'location', 'farmer', 'buyer'
    )
    
    # Apply filters
    if firm:
        submissions_qs = submissions_qs.filter(firm=firm)
    
    if category_id:
        try:
            category = ProductCategory.objects.get(id=category_id)
            submissions_qs = submissions_qs.filter(category=category)
        except ProductCategory.DoesNotExist:
            pass
    
    if product_id:
        try:
            product = Product.objects.get(id=product_id)
            submissions_qs = submissions_qs.filter(product=product)
        except Product.DoesNotExist:
            pass
    
    if location_id:
        try:
            location = Location.objects.get(id=location_id)
            submissions_qs = submissions_qs.filter(location=location)
        except Location.DoesNotExist:
            pass
    
    # Apply user permissions
    if not user_profile.is_admin():
        if user_profile.firm:
            submissions_qs = submissions_qs.filter(firm=user_profile.firm)
        
        if user_profile.is_category_head() and hasattr(user_profile, 'category_head_profile'):
            scoped_categories = user_profile.category_head_profile.scoped_categories.all()
            if scoped_categories:
                submissions_qs = submissions_qs.filter(category__in=scoped_categories)
    
    # Add highlighting
    highlighted_submissions = ReportingService.get_price_submissions_with_highlights(
        submissions_qs, group_by_location=True
    )
    
    # Handle export requests
    if export_format in ['csv', 'xlsx']:
        return export_submissions_data(
            highlighted_submissions, 
            export_format, 
            f'firm_daily_report_{report_date.isoformat()}'
        )
    
    # Get filter choices for the form
    firms = Firm.objects.all()
    categories = ProductCategory.objects.all()
    products = Product.objects.all()
    locations = Location.objects.all()
    
    # Limit choices based on user permissions
    if not user_profile.is_admin():
        if user_profile.firm:
            firms = firms.filter(id=user_profile.firm.id)
        
        if user_profile.is_category_head() and hasattr(user_profile, 'category_head_profile'):
            scoped_categories = user_profile.category_head_profile.scoped_categories.all()
            if scoped_categories:
                categories = scoped_categories
                products = products.filter(category__in=scoped_categories)
    
    context = {
        'user_profile': user_profile,
        'firm_reports': firm_reports,
        'submissions': highlighted_submissions,
        'report_date': report_date,
        'selected_firm': firm,
        'firms': firms,
        'categories': categories,
        'products': products,
        'locations': locations,
        'filters': {
            'firm': firm_id,
            'category': category_id,
            'product': product_id,
            'location': location_id
        }
    }
    
    return render(request, 'core/reports/firm_daily.html', context)


@login_required
def category_wise_report(request):
    """Category-wise report with trends"""
    user_profile = request.user.profile
    
    # Get date range
    date_to = timezone.now().date()
    date_from = date_to - timedelta(days=30)
    
    if request.GET.get('date_from'):
        try:
            date_from = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('date_to'):
        try:
            date_to = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    firm = None
    if request.GET.get('firm') and user_profile.is_admin():
        try:
            firm = Firm.objects.get(id=request.GET.get('firm'))
        except Firm.DoesNotExist:
            pass
    elif not user_profile.is_admin():
        firm = user_profile.firm
    
    # Get the report
    report_data = ReportingService.get_category_wise_report(
        date_from=date_from,
        date_to=date_to,
        firm=firm
    )
    
    context = {
        'user_profile': user_profile,
        'report_data': report_data,
        'date_from': date_from,
        'date_to': date_to,
        'selected_firm': firm,
        'firms': Firm.objects.all() if user_profile.is_admin() else []
    }
    
    return render(request, 'core/reports/category_wise.html', context)


@login_required
def product_trends_report(request):
    """Product price trends report"""
    user_profile = request.user.profile
    
    product_id = request.GET.get('product')
    days = int(request.GET.get('days', 30))
    
    if not product_id:
        messages.error(request, "Please select a product to view trends.")
        return redirect('core:reports_index')
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        messages.error(request, "Product not found.")
        return redirect('core:reports_index')
    
    # Check permissions
    if not user_profile.is_admin():
        if user_profile.is_category_head() and hasattr(user_profile, 'category_head_profile'):
            scoped_categories = user_profile.category_head_profile.scoped_categories.all()
            if scoped_categories and product.category not in scoped_categories:
                messages.error(request, "You don't have permission to view this product.")
                return redirect('core:reports_index')
    
    firm = None
    if not user_profile.is_admin():
        firm = user_profile.firm
    elif request.GET.get('firm'):
        try:
            firm = Firm.objects.get(id=request.GET.get('firm'))
        except Firm.DoesNotExist:
            pass
    
    # Get trends data
    trends = ReportingService.get_product_price_trends(
        product_id=product.id,
        days=days,
        firm=firm
    )
    
    # Convert to format suitable for charts
    chart_data = {
        'labels': [trend['date'].isoformat() for trend in trends],
        'datasets': [
            {
                'label': 'Average Price',
                'data': [float(trend['avg_price'] or 0) for trend in trends],
                'borderColor': 'rgb(59, 130, 246)',
                'backgroundColor': 'rgba(59, 130, 246, 0.1)'
            },
            {
                'label': 'Min Price',
                'data': [float(trend['min_price'] or 0) for trend in trends],
                'borderColor': 'rgb(34, 197, 94)',
                'backgroundColor': 'rgba(34, 197, 94, 0.1)'
            },
            {
                'label': 'Max Price',
                'data': [float(trend['max_price'] or 0) for trend in trends],
                'borderColor': 'rgb(239, 68, 68)',
                'backgroundColor': 'rgba(239, 68, 68, 0.1)'
            }
        ]
    }
    
    context = {
        'user_profile': user_profile,
        'product': product,
        'trends': trends,
        'chart_data': json.dumps(chart_data),
        'days': days,
        'selected_firm': firm,
        'firms': Firm.objects.all() if user_profile.is_admin() else []
    }
    
    return render(request, 'core/reports/product_trends.html', context)


def export_submissions_data(queryset, format_type, filename):
    """Export submissions data to CSV or Excel"""
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Firm', 'Category', 'Product', 'Location', 'Farmer',
            'Price per Unit', 'Quantity', 'Unit', 'Total Value',
            'Buyer', 'Status', 'Notes', 'Approved By', 'Approved At'
        ])
        
        for submission in queryset:
            writer.writerow([
                submission.date,
                submission.firm.name,
                submission.category.name,
                submission.product.name,
                submission.location.name,
                submission.farmer.name,
                submission.price_per_unit,
                submission.quantity,
                submission.unit,
                submission.total_value,
                submission.buyer.get_full_name() or submission.buyer.username,
                submission.get_status_display(),
                submission.notes,
                submission.approved_by.get_full_name() if submission.approved_by else '',
                submission.approved_at.strftime('%Y-%m-%d %H:%M') if submission.approved_at else ''
            ])
        
        return response
    
    elif format_type == 'xlsx' and OPENPYXL_AVAILABLE:
        import io
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Submissions'
        
        # Headers
        headers = [
            'Date', 'Firm', 'Category', 'Product', 'Location', 'Farmer',
            'Price per Unit', 'Quantity', 'Unit', 'Total Value',
            'Buyer', 'Status', 'Notes', 'Approved By', 'Approved At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # Data
        for row_num, submission in enumerate(queryset, 2):
            data = [
                submission.date,
                submission.firm.name,
                submission.category.name,
                submission.product.name,
                submission.location.name,
                submission.farmer.name,
                float(submission.price_per_unit),
                float(submission.quantity),
                submission.unit,
                float(submission.total_value),
                submission.buyer.get_full_name() or submission.buyer.username,
                submission.get_status_display(),
                submission.notes,
                submission.approved_by.get_full_name() if submission.approved_by else '',
                submission.approved_at.strftime('%Y-%m-%d %H:%M') if submission.approved_at else ''
            ]
            
            for col_num, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)
        
        # Save to BytesIO
        virtual_file = io.BytesIO()
        workbook.save(virtual_file)
        virtual_file.seek(0)
        
        response.write(virtual_file.getvalue())
        return response
    
    else:
        # Fallback to CSV if Excel is not available
        return export_submissions_data(queryset, 'csv', filename)